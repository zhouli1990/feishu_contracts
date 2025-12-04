from __future__ import annotations
import argparse
import json
import math
import os
import time
import logging
from typing import Any, Dict, List, Optional

import pandas as pd
import yaml
from openpyxl import load_workbook

logger = logging.getLogger("transform")


LOOKUP_TABLES: Dict[str, Dict[Any, Any]] = {}


def _to_str(x: Any) -> str:
    if x is None or (isinstance(x, float) and math.isnan(x)):
        return ""
    return str(x)


def _java_dt_to_py(pattern: str) -> str:
    mapping = {"yyyy": "%Y", "MM": "%m", "dd": "%d", "HH": "%H", "mm": "%M", "ss": "%S"}
    out = pattern
    for k, v in mapping.items():
        out = out.replace(k, v)
    return out


def _try_parse_date(txt: str, fmts: List[str]) -> Optional[pd.Timestamp]:
    if not isinstance(txt, str):
        return None
    s = txt.strip()
    if s == "":
        return None
    numeric = s.lstrip("+-")
    numeric_core = numeric.split(".", 1)[0]
    if numeric and numeric.replace(".", "", 1).isdigit() and len(numeric_core) >= 10:
        try:
            if "." in numeric:
                value: float | int = float(s)
            else:
                value = int(s)
            digits = len(numeric_core)
            if digits >= 19:
                unit = "ns"
            elif digits >= 16:
                unit = "us"
            elif digits >= 13:
                unit = "ms"
            else:
                unit = "s"
            return pd.to_datetime(value, unit=unit, errors="raise")
        except Exception:
            pass
    for f in fmts:
        try:
            return pd.to_datetime(s, format=_java_dt_to_py(f), errors="raise")
        except Exception:
            continue
    try:
        return pd.to_datetime(s, errors="raise")
    except Exception:
        return None


def _number_parse(txt: Any, thousands: str = ",", decimal: str = ".") -> Optional[float]:
    if txt is None:
        return None
    if isinstance(txt, (int, float)) and not isinstance(txt, bool):
        return float(txt)
    s = str(txt).strip()
    if s == "":
        return None
    if thousands:
        s = s.replace(thousands, "")
    if decimal != ".":
        s = s.replace(decimal, ".")
    try:
        return float(s)
    except Exception:
        return None


def _dict_lookup_value(table: Dict[Any, Any], key: Any) -> Optional[Any]:
    if not table:
        return None
    candidates: List[Any] = []
    seen = set()

    def _push(candidate: Any):
        if candidate in seen:
            return
        seen.add(candidate)
        candidates.append(candidate)

    _push(key)
    if isinstance(key, str):
        stripped = key.strip()
        if stripped != key:
            _push(stripped)
        if stripped:
            try:
                _push(int(stripped))
            except Exception:
                pass
    else:
        try:
            _push(str(key))
        except Exception:
            pass

    for candidate in candidates:
        if candidate in table:
            return table[candidate]
    return None


# ---------------- transforms ----------------

def tf_trim(values: List[Any], _):
    return [(_to_str(v).strip() if v is not None else None) for v in values]


def tf_json_parse(values: List[Any], _):
    out = []
    for v in values:
        if isinstance(v, str):
            s = v.strip()
            if s == "":
                out.append("")
                continue
            try:
                out.append(json.loads(s))
            except Exception:
                out.append(s)
        else:
            out.append(v)
    return out


def tf_form_pick(values: List[Any], params: Dict[str, Any]):
    field = params.get("field")
    field_pairs = params.get("field_pairs")
    unique = params.get("unique", True)
    res: List[Any] = []
    for v in values:
        seq = v if isinstance(v, list) else [v]
        for item in seq:
            if field_pairs and isinstance(item, dict):
                res.append({k: item.get(k) for k in field_pairs})
            elif field and isinstance(item, dict):
                res.append(item.get(field))
            else:
                res.append(item)
    if unique:
        seen = set()
        tmp = []
        for x in res:
            key = json.dumps(x, ensure_ascii=False, sort_keys=True) if isinstance(x, dict) else x
            if key in seen:
                continue
            seen.add(key)
            tmp.append(x)
        res = tmp
    return res


def tf_format_each(values: List[Any], params: Dict[str, Any]):
    tpl = params.get("template", "{x}")
    out: List[str] = []
    for v in values:
        if isinstance(v, dict):
            try:
                out.append(tpl.format(**{k: _to_str(v.get(k)) for k in v.keys()}))
            except Exception:
                out.append(str(v))
        else:
            try:
                out.append(tpl.format(x=_to_str(v)))
            except Exception:
                out.append(_to_str(v))
    return out


def tf_join_agg(values: List[Any], params: Dict[str, Any]):
    sep = params.get("sep", ", ")
    unique = params.get("unique", True)
    flat: List[str] = []
    for v in values:
        if isinstance(v, list):
            flat.extend([_to_str(i) for i in v])
        else:
            flat.append(_to_str(v))
    if unique:
        seen = set()
        tmp = []
        for s in flat:
            if s in seen:
                continue
            seen.add(s)
            tmp.append(s)
        flat = tmp
    return [sep.join([s for s in flat if s != ""]) if flat else ""]


def tf_number_parse(values: List[Any], params: Dict[str, Any]):
    thousands = params.get("thousands", ",")
    decimal = params.get("decimal", ".")
    return [_number_parse(v, thousands, decimal) for v in values]


def tf_round(values: List[Any], params: Dict[str, Any]):
    scale = int(params) if isinstance(params, int) else int(params.get("", 2)) if isinstance(params, dict) else 2
    out = []
    for v in values:
        if isinstance(v, (int, float)):
            out.append(round(float(v), scale))
        else:
            out.append(v)
    return out


def tf_date_parse(values: List[Any], params: Dict[str, Any]):
    fmts = params.get("input_formats", [])
    out = []
    for v in values:
        ts = _try_parse_date(_to_str(v), fmts)
        out.append(ts)
    return out


def tf_date_format(values: List[Any], params: Dict[str, Any]):
    pattern = params if isinstance(params, str) else params.get("", None) if isinstance(params, dict) else None
    if pattern is None and isinstance(params, dict):
        pattern = params.get("pattern")
    pyfmt = _java_dt_to_py(pattern or "%Y-%m-%d")
    out = []
    for v in values:
        if isinstance(v, pd.Timestamp):
            out.append(v.strftime(pyfmt))
        elif isinstance(v, str) and v:
            try:
                out.append(pd.to_datetime(v).strftime(pyfmt))
            except Exception:
                out.append(v)
        else:
            out.append("")
    return out


def tf_to_value_label(values: List[Any], params: Dict[str, Any]) -> List[Any]:
    value_field = params.get("value_field")
    label_field = params.get("label_field")
    table_name = params.get("value_dict")
    keep_original = params.get("keep_original", True)
    default = params.get("default", "")
    unique = params.get("unique", True)

    table = LOOKUP_TABLES.get(table_name) if table_name else None
    res: List[Any] = []
    for v in values:
        seq = v if isinstance(v, list) else [v]
        for item in seq:
            if isinstance(item, dict):
                raw_val = item.get(value_field)
                mapped = None
                if table is not None:
                    mapped = _dict_lookup_value(table, raw_val)
                if mapped is not None:
                    final_val = mapped
                elif keep_original and raw_val not in (None, ""):
                    final_val = raw_val
                else:
                    final_val = default
                label_val = item.get(label_field)
                res.append({"value": final_val, "label": label_val})
            else:
                if item not in (None, ""):
                    res.append({"value": item, "label": None})
    if unique:
        seen = set(); tmp: List[Any] = []
        for x in res:
            key = json.dumps(x, ensure_ascii=False, sort_keys=True)
            if key in seen:
                continue
            seen.add(key)
            tmp.append(x)
        res = tmp
    return res


def tf_json_stringify(values: List[Any], _params: Dict[str, Any]) -> List[Any]:
    try:
        return [json.dumps(values, ensure_ascii=False)]
    except Exception:
        return [str(values)]


def tf_dict_lookup(values: List[Any], params: Any) -> List[Any]:
    if isinstance(params, str):
        table_name = params
        default = ""
        keep_original = True
    elif isinstance(params, dict):
        table_name = params.get("table")
        default = params.get("default", "")
        keep_original = params.get("keep_original", True)
    else:
        return list(values)

    if not table_name:
        return list(values)

    table = LOOKUP_TABLES.get(table_name, {})
    out: List[Any] = []
    for v in values:
        mapped = _dict_lookup_value(table, v)
        if mapped is not None:
            out.append(mapped)
        elif keep_original and v not in (None, ""):
            out.append(v)
        else:
            out.append(default)
    return out


TRANSFORMS = {
    "trim": tf_trim,
    "json_parse": tf_json_parse,
    "form_pick": tf_form_pick,
    "form_picker_names": lambda v, p: tf_form_pick(v, {"field": "name", **(p or {})}),
    "format_each": tf_format_each,
    "join_agg": tf_join_agg,
    "number_parse": tf_number_parse,
    "round": tf_round,
    "date_parse": tf_date_parse,
    "date_format": tf_date_format,
    "to_value_label": tf_to_value_label,
    "json_stringify": tf_json_stringify,
    "dict": tf_dict_lookup,
}


def apply_chain(values: List[Any], chain: List[Any]) -> List[Any]:
    out = list(values)
    for step in chain or []:
        if isinstance(step, str):
            fn = TRANSFORMS.get(step)
            params = {}
        elif isinstance(step, dict):
            k = next(iter(step))
            fn = TRANSFORMS.get(k)
            params = step[k]
        else:
            continue
        if fn is None:
            continue
        out = fn(out, params)
    return out


def read_excel_all(path: str) -> Dict[str, pd.DataFrame]:
    x = pd.read_excel(path, sheet_name=None, dtype=str)
    for k in x:
        x[k] = x[k].fillna("")
    return x


def header_index(ws) -> Dict[str, int]:
    headers = {}
    for j, cell in enumerate(ws[1], start=1):
        headers[_to_str(cell.value).strip()] = j
    return headers


def get_values_from_source(df_map: Dict[str, pd.DataFrame], base_row: Dict[str, Any], frm: Dict[str, Any], where: Optional[Dict[str, Any]], join_key: str, groups: Optional[Dict[str, Any]] = None) -> List[Any]:
    sheet = frm["sheet"]
    col = frm["column"]
    if sheet not in df_map:
        return []
    if sheet == base_row.get("__base_sheet__"):
        return [base_row.get(col, "")]
    join_val = base_row.get(join_key, "")
    base_df = df_map[sheet]
    if groups and sheet in groups:
        grp = groups[sheet]
        try:
            q = grp.get_group(join_val)
        except Exception:
            q = base_df.iloc[0:0]
    else:
        q = base_df[base_df.get(join_key, pd.Series([""] * len(base_df))) == join_val]
    if where:
        for k, v in where.items():
            if k in q.columns:
                q = q[q[k] == str(v)]
    if col in q.columns:
        return q[col].tolist()
    return []


def process_one_to_one(ts_spec: Dict[str, Any], df_map: Dict[str, pd.DataFrame], wb, join_key: str, groups: Optional[Dict[str, Any]] = None) -> int:
    """通用的一对一目标表写入：以 details 为基表（如存在），否则以 ts_spec.source 为基表，
    按 join_key 维度聚合同一合同号下其他表的 where 命中行，并执行 transform 链得到单值输出。
    """
    ws = wb[ts_spec["name"]]
    hdr = header_index(ws)
    base_sheet = "details" if "details" in df_map else ts_spec.get("source")
    base_df = df_map[base_sheet]
    out_rows: List[Dict[str, Any]] = []
    total = len(base_df)
    for i, (_, r) in enumerate(base_df.iterrows(), start=1):
        base = r.to_dict()
        base["__base_sheet__"] = base_sheet
        row_out: Dict[str, Any] = {}
        for mp in ts_spec.get("mappings", []):
            to_col = mp["to"]["column"]
            fr = mp.get("from", [])
            where = mp.get("where")
            vals: List[Any] = []
            if not fr:
                vals = []
            else:
                for f in fr:
                    vals.extend(get_values_from_source(df_map, base, f, where, join_key, groups))
            vals = apply_chain(vals, mp.get("transform", []))
            final = vals[0] if vals else mp.get("default", "")
            row_out[to_col] = final
        out_rows.append(row_out)
        if i % 1000 == 0:
            logger.info("progress sheet=%s %d/%d", ts_spec["name"], i, total, extra={"is_progress": True})
    # 追加写入
    start_row = ws.max_row + 1 if ws.max_row >= 1 else 2
    for row in out_rows:
        i = start_row
        for k, v in row.items():
            if k in hdr:
                ws.cell(row=i, column=hdr[k], value=v)
        start_row += 1
    return len(out_rows)


def process_simple_append(ts_spec: Dict[str, Any], df_map: Dict[str, pd.DataFrame], wb) -> int:
    ws = wb[ts_spec["name"]]
    hdr = header_index(ws)
    src_name = ts_spec.get("source")
    if not src_name or src_name not in df_map:
        # 源 Sheet 缺失：跳过该目标写入（常见于可选子表：payments 等）
        return 0
    src = df_map[src_name]
    start_row = ws.max_row + 1 if ws.max_row >= 1 else 2
    count = 0
    for _, r in src.iterrows():
        base = r.to_dict(); base["__base_sheet__"] = ts_spec.get("source")
        row_out: Dict[str, Any] = {}
        for mp in ts_spec["mappings"]:
            to_col = mp["to"]["column"]
            fr = mp.get("from", [])
            where = mp.get("where")
            vals: List[Any] = []
            if not fr:
                vals = []
            else:
                for f in fr:
                    vals.extend([base.get(f["column"], "")] if f["sheet"] == ts_spec.get("source") else [])
            vals = apply_chain(vals, mp.get("transform", []))
            final = vals[0] if vals else (mp.get("default", ""))
            row_out[to_col] = final
        # write
        for k, v in row_out.items():
            if k in hdr:
                ws.cell(row=start_row, column=hdr[k], value=v)
        start_row += 1
        count += 1
    return count


def run(source: str, template: str, mapping_path: str, out_path: str):
    with open(mapping_path, "r", encoding="utf-8") as f:
        mapping = yaml.safe_load(f)
    join_key = next(iter(mapping.get("join", {}).get("keys", {})))  # e.g., contract_number
    global LOOKUP_TABLES
    LOOKUP_TABLES = mapping.get("dict", {}) or {}
    for ds in (mapping.get("dict_sources") or []):
        try:
            name = ds.get("name"); path = ds.get("path")
            sheet = ds.get("sheet")
            key_col = ds.get("key_column"); val_col = ds.get("value_column")
            if not name or not path or not key_col or not val_col:
                continue
            if not os.path.isabs(path):
                base_dir = os.path.dirname(mapping_path)
                path = os.path.normpath(os.path.join(base_dir, path))
            df = pd.read_excel(path, sheet_name=(sheet if sheet else 0), dtype=str)
            if isinstance(df, dict):
                if sheet and sheet in df:
                    df = df[sheet]
                else:
                    df = list(df.values())[0]
            df = df.fillna("")
            table: Dict[Any, Any] = {}
            for _, r in df.iterrows():
                k = str(r.get(key_col, "")).strip()
                v = str(r.get(val_col, "")).strip()
                if k != "":
                    table[k] = v
            LOOKUP_TABLES[name] = table
            logger.info("loaded dict source name=%s path=%s size=%d", name, path, len(table))
        except Exception as e:
            logger.warning("failed to load dict source name=%s path=%s error=%s", ds.get("name"), ds.get("path"), e)
    df_map = read_excel_all(source)
    wb = load_workbook(template)

    # 逐目标表处理：one_to_one 使用通用聚合；one_to_many 逐行追加
    logger.info("start transform source=%s template=%s mapping=%s out=%s", source, template, mapping_path, out_path, extra={"is_progress": True})
    used_sheets = set()
    form_attr_names = set()
    for ts in mapping.get("target_sheets", []):
        for mp in ts.get("mappings", []):
            fr = mp.get("from", [])
            for f in fr:
                s = f.get("sheet")
                if s:
                    used_sheets.add(s)
                    if s == "form":
                        w = mp.get("where", {})
                        if isinstance(w, dict):
                            attr = w.get("attribute_name")
                            if isinstance(attr, str) and attr:
                                form_attr_names.add(attr)
    if "form" in df_map and form_attr_names and "attribute_name" in df_map["form"].columns:
        df_map["form"] = df_map["form"][df_map["form"]["attribute_name"].isin(list(form_attr_names))]
    groups: Dict[str, Any] = {}
    for s in used_sheets:
        df = df_map.get(s)
        if df is None:
            continue
        if join_key in df.columns:
            groups[s] = df.groupby(join_key)
    for ts in mapping["target_sheets"]:
        name = ts["name"]
        if name not in wb.sheetnames:
            continue
        policy = ts.get("row_policy", "one_to_one")
        logger.info("begin sheet=%s policy=%s", name, policy, extra={"is_progress": True})
        t0 = time.time()
        if policy == "one_to_one":
            rows = process_one_to_one(ts, df_map, wb, join_key, groups)
        else:
            rows = process_simple_append(ts, df_map, wb)
        elapsed = time.time() - t0
        logger.info("sheet=%s policy=%s rows=%d elapsed=%.1fs", name, policy, int(rows), elapsed, extra={"is_progress": True})

    wb.save(out_path)
    logger.info("done out=%s", out_path, extra={"is_progress": True})


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--source", required=True)
    p.add_argument("--template", required=True)
    p.add_argument("--mapping", required=False, default=os.path.join(os.path.dirname(__file__), "mapping.yaml"))
    p.add_argument("--out", required=True)
    args = p.parse_args()
    run(args.source, args.template, args.mapping, args.out)


if __name__ == "__main__":
    main()
